#!/usr/bin/env python3
"""
Application Web de Lettrage Comptable Automatique
==================================================
Interface Streamlit pour le lettrage comptable professionnel

Auteur: Système de Lettrage Automatique
Version: 1.0
Date: Décembre 2025
"""

import streamlit as st
import pandas as pd
import numpy as np
from io import BytesIO
from datetime import datetime
from collections import defaultdict
import sys
import os
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter


# Import des modules de lettrage
# (Le code sera intégré directement dans l'application)

# ============================================================================
# CLASSE DE LETTRAGE (Intégrée)
# ============================================================================

class LettrageComptableWeb:
    """Classe de lettrage adaptée pour Streamlit"""
    
    def __init__(self, df_reglement, df_justif):
        self.df_reglement = df_reglement
        self.df_justif = df_justif
        self.codes_lettrage = {}
        self.compteur_pfa = 0
        self.compteur_preg = 0
        
        self.stats = {
            'total_transactions': 0,
            'nb_reglements': 0,
            'nb_factures': 0,
            'nb_lettres_a': 0,
            'nb_pfa': 0,
            'nb_preg': 0,
            'nb_non_reglee': 0,
            'nb_non_lettre': 0
        }
    
    def identifier_type_transaction(self, row):
        """Identifie si une transaction est un règlement ou une facture"""
        solde = row.get('SOLDE', 0)
        if pd.isna(solde):
            solde = 0
        
        if solde > 0:
            return 'REGLEMENT'
        elif solde < 0:
            return 'FACTURE'
        else:
            debit = row.get(' MONTANT_DEBIT (ANNULATION FT OU AVOIR OU REGLEMENT)', 0)
            credit = row.get('MONTANT_CREDIT (FACTURE OU REGLEMENT ANNULE)', 0)
            
            if pd.notna(debit) and debit > 0:
                return 'REGLEMENT'
            elif pd.notna(credit) and credit > 0:
                return 'FACTURE'
            else:
                return 'INDETERMINE'
    
    def construire_mapping_reglement_factures(self):
        """Construit les dictionnaires de correspondance"""
        reglement_vers_factures = defaultdict(list)
        facture_vers_reglements = defaultdict(list)
        
        col_num_reglement = 'Numéro Règlement'
        col_num_ap_facture = 'Num AP Facture'
        
        for _, row in self.df_reglement.iterrows():
            num_reglement = row.get(col_num_reglement)
            num_ap_facture = row.get(col_num_ap_facture)
            
            if pd.notna(num_reglement):
                try:
                    num_reglement = int(float(num_reglement))
                except:
                    pass
            
            if pd.notna(num_ap_facture):
                try:
                    num_ap_facture = int(float(num_ap_facture))
                except:
                    pass
            
            if pd.notna(num_reglement) and pd.notna(num_ap_facture):
                reglement_vers_factures[num_reglement].append(num_ap_facture)
                facture_vers_reglements[num_ap_facture].append(num_reglement)
        
        return reglement_vers_factures, facture_vers_reglements
    
    def trouver_montant_transaction(self, identifiant, type_trans):
        """Trouve le montant d'une transaction"""
        if type_trans == 'FACTURE':
            mask = (self.df_justif['NUM_AP'] == identifiant)
        else:
            mask = (self.df_justif['TRANSACTION_NUMBER'] == identifiant)
        
        if not mask.any():
            if type_trans == 'FACTURE':
                mask = (self.df_justif['NUM_AP'].astype(str) == str(identifiant))
            else:
                mask = (self.df_justif['TRANSACTION_NUMBER'].astype(str) == str(identifiant))
        
        if mask.any():
            solde = self.df_justif.loc[mask, 'SOLDE'].iloc[0]
            return abs(solde) if pd.notna(solde) else 0
        
        return 0
    
    def normaliser_identifiant(self, identifiant):
        """Normalise un identifiant"""
        if pd.isna(identifiant) or identifiant == '(vide)':
            return None
        try:
            return int(float(identifiant))
        except:
            return str(identifiant)
    
    def effectuer_lettrage(self):
        """Effectue le lettrage complet"""
        reg_to_fact, fact_to_reg = self.construire_mapping_reglement_factures()
        
        # Identifier les types
        self.df_justif['TYPE_TRANSACTION'] = self.df_justif.apply(
            self.identifier_type_transaction, axis=1
        )
        
        self.stats['nb_reglements'] = (self.df_justif['TYPE_TRANSACTION'] == 'REGLEMENT').sum()
        self.stats['nb_factures'] = (self.df_justif['TYPE_TRANSACTION'] == 'FACTURE').sum()
        
        # Traiter les groupes
        transactions_traitees = set()
        
        # Traiter les règlements
        for idx, row in self.df_justif.iterrows():
            type_trans = row['TYPE_TRANSACTION']
            
            if type_trans != 'REGLEMENT':
                continue
            
            identifiant = self.normaliser_identifiant(row.get('TRANSACTION_NUMBER'))
            if identifiant is None or identifiant in transactions_traitees:
                continue
            
            factures_liees = reg_to_fact.get(identifiant, [])
            
            if len(factures_liees) == 0:
                self.codes_lettrage[identifiant] = 'NON LETTRE'
                transactions_traitees.add(identifiant)
                continue
            
            # Construire le groupe complet
            tous_reglements = set([identifiant])
            for fact_id in factures_liees:
                reglements_de_facture = fact_to_reg.get(fact_id, [])
                tous_reglements.update(reglements_de_facture)
            
            toutes_factures = set(factures_liees)
            for reg_id in tous_reglements:
                factures_de_reglement = reg_to_fact.get(reg_id, [])
                toutes_factures.update(factures_de_reglement)
            
            if any(t in transactions_traitees for t in tous_reglements) or \
               any(t in transactions_traitees for t in toutes_factures):
                continue
            
            # Calculer les totaux
            total_reglements = sum(self.trouver_montant_transaction(r, 'REGLEMENT') for r in tous_reglements)
            total_factures = sum(self.trouver_montant_transaction(f, 'FACTURE') for f in toutes_factures)
            
            ecart = total_reglements - total_factures
            
            # Déterminer le code
            if abs(ecart) <= 0.001:
                code = 'A'
            elif ecart > 0:
                self.compteur_pfa += 1
                code = f'P(FA){self.compteur_pfa}'
            else:
                self.compteur_preg += 1
                code = f'P(REG){self.compteur_preg}'
            
            # Appliquer le code
            for reg_id in tous_reglements:
                self.codes_lettrage[reg_id] = code
                transactions_traitees.add(reg_id)
            
            for fact_id in toutes_factures:
                self.codes_lettrage[fact_id] = code
                transactions_traitees.add(fact_id)
        
        # Traiter les factures non réglées
        for idx, row in self.df_justif.iterrows():
            type_trans = row['TYPE_TRANSACTION']
            
            if type_trans != 'FACTURE':
                continue
            
            identifiant = row.get('NUM_AP')
            if pd.isna(identifiant) or identifiant == '(vide)':
                identifiant = row.get('TRANSACTION_NUMBER')
            
            identifiant = self.normaliser_identifiant(identifiant)
            if identifiant is None or identifiant in transactions_traitees:
                continue
            
            self.codes_lettrage[identifiant] = 'NON REGLEE'
            transactions_traitees.add(identifiant)
        
        # Appliquer les codes
        self.df_justif['LETTRAGE_AUTO'] = self.df_justif.apply(
            lambda row: self.obtenir_code_pour_ligne(row), axis=1
        )
        
        # Calculer les statistiques
        self.stats['total_transactions'] = len(self.df_justif)
        self.stats['nb_lettres_a'] = (self.df_justif['LETTRAGE_AUTO'] == 'A').sum()
        self.stats['nb_non_reglee'] = (self.df_justif['LETTRAGE_AUTO'] == 'NON REGLEE').sum()
        self.stats['nb_non_lettre'] = (self.df_justif['LETTRAGE_AUTO'] == 'NON LETTRE').sum()
        self.stats['nb_pfa'] = self.df_justif['LETTRAGE_AUTO'].str.match(r'P\(FA\)\d+', na=False).sum()
        self.stats['nb_preg'] = self.df_justif['LETTRAGE_AUTO'].str.match(r'P\(REG\)\d+', na=False).sum()
        
        return self.df_justif, self.stats
    
    def obtenir_code_pour_ligne(self, row):
        """Obtient le code pour une ligne"""
        type_trans = row['TYPE_TRANSACTION']
        
        if type_trans == 'INDETERMINE':
            return 'INDETERMINE'
        
        if type_trans == 'FACTURE':
            identifiant = row.get('NUM_AP')
            if pd.isna(identifiant) or identifiant == '(vide)':
                identifiant = row.get('TRANSACTION_NUMBER')
        else:
            identifiant = row.get('TRANSACTION_NUMBER')
        
        identifiant = self.normaliser_identifiant(identifiant)
        
        return self.codes_lettrage.get(identifiant, 'NON LETTRE')


# ============================================================================
# CONFIGURATION STREAMLIT
# ============================================================================

st.set_page_config(
    page_title="Lettrage Comptable Automatique ",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded"
)

# CSS personnalisé
st.markdown("""
<style>
    .main-header {
        font-size: 2.5rem;
        color: #1f77b4;
        text-align: center;
        padding: 1rem 0;
        border-bottom: 3px solid #1f77b4;
        margin-bottom: 2rem;
    }
    .metric-card {
        background-color: #f0f2f6;
        padding: 1.5rem;
        border-radius: 0.5rem;
        border-left: 5px solid #1f77b4;
    }
    .success-box {
        background-color: #d4edda;
        border-left: 5px solid #28a745;
        padding: 1rem;
        border-radius: 0.3rem;
        margin: 1rem 0;
    }
    .warning-box {
        background-color: #fff3cd;
        border-left: 5px solid #ffc107;
        padding: 1rem;
        border-radius: 0.3rem;
        margin: 1rem 0;
    }
    .info-box {
        background-color: #d1ecf1;
        border-left: 5px solid #17a2b8;
        padding: 1rem;
        border-radius: 0.3rem;
        margin: 1rem 0;
    }
</style>
""", unsafe_allow_html=True)


# ============================================================================
# INTERFACE PRINCIPALE
# ============================================================================

def main():
    # En-tête
    st.markdown('<h1 class="main-header">📊 Lettrage Comptable Automatique</h1>', unsafe_allow_html=True)
    
    # Barre latérale
    with st.sidebar:
        st.image("https://via.placeholder.com/300x100/1f77b4/ffffff?text=Lettrage+2022", use_container_width=True)
        st.markdown("---")
        st.markdown("### 📋 Instructions")
        st.markdown("""
        1. Téléchargez **Règlement Global**
        2. Téléchargez **Justificatif Global**
        3. Cliquez sur **Lancer le Lettrage**
        4. Téléchargez les résultats
        """)
        st.markdown("---")
        st.markdown("### 🎯 Codes de Lettrage")
        st.markdown("""
        - **A** : Lettré exactement
        - **P(FA)i** : Règlements > Factures
        - **P(REG)i** : Règlements < Factures
        - **NON REGLEE** : Factures non payées
        - **NON LETTRE** : Règlements orphelins
        """)
        st.markdown("---")
        st.markdown("### ℹ️ Version")
        st.info("Version 1.0 - Décembre 2025")
    
    # Onglets principaux
    tab1, tab2, tab3 = st.tabs(["📁 Upload & Lettrage", "📊 Statistiques", "📖 Documentation"])
    
    # ========== ONGLET 1: UPLOAD & LETTRAGE ==========
    with tab1:
        st.markdown("## 📤 Téléchargement des Fichiers")
        
        col1, col2 = st.columns(2)
        
        with col1:
            st.markdown("### 1️⃣ Règlement Global")
            fichier_reglement = st.file_uploader(
                "Fichier Excel contenant les correspondances règlements ↔ factures",
                type=['xlsx', 'xls'],
                key="reglement"
            )
            if fichier_reglement:
                st.success(f"✅ Fichier chargé: {fichier_reglement.name}")
        
        with col2:
            st.markdown("### 2️⃣ Justificatif Global")
            fichier_justif = st.file_uploader(
                "Fichier Excel avec toutes les transactions à lettrer",
                type=['xlsx', 'xls'],
                key="justif"
            )
            if fichier_justif:
                st.success(f"✅ Fichier chargé: {fichier_justif.name}")
        
        st.markdown("---")
        
        # Bouton de lettrage
        if fichier_reglement and fichier_justif:
            if st.button("🚀 Lancer le Lettrage", type="primary", use_container_width=True):
                with st.spinner("⏳ Lettrage en cours..."):
                    try:
                        # Charger les fichiers
                        df_reglement_raw = pd.read_excel(fichier_reglement, header=None)
                        
                        # Trouver l'en-tête
                        header_row = None
                        for idx, row in df_reglement_raw.iterrows():
                            row_values = [str(val).lower() for val in row if pd.notna(val)]
                            if any('num' in val and 'règlement' in val for val in row_values):
                                header_row = idx
                                break
                        
                        if header_row is None:
                            st.error("❌ Impossible de trouver l'en-tête dans le fichier règlement")
                            return
                        
                        df_reglement = pd.read_excel(fichier_reglement, header=header_row)
                        df_justif = pd.read_excel(fichier_justif)
                        
                        # Nettoyer
                        df_justif = df_justif[~df_justif['Étiquettes de lignes'].str.contains('Total', na=False)]
                        
                        # Effectuer le lettrage
                        lettrage = LettrageComptableWeb(df_reglement, df_justif)
                        df_resultat, stats = lettrage.effectuer_lettrage()
                        
                        # Stocker dans session_state
                        st.session_state['df_resultat'] = df_resultat
                        st.session_state['stats'] = stats
                        st.session_state['lettrage_effectue'] = True
                        
                        st.success("✅ Lettrage terminé avec succès !")
                        st.balloons()
                        
                    except Exception as e:
                        st.error(f"❌ Erreur lors du lettrage: {str(e)}")
                        st.exception(e)
        
        # Affichage des résultats
        if st.session_state.get('lettrage_effectue', False):
            st.markdown("---")
            st.markdown("## 📊 Résultats du Lettrage")
            
            stats = st.session_state['stats']
            df_resultat = st.session_state['df_resultat']
            
            # Métriques principales
            col1, col2, col3, col4, col5 = st.columns(5)
            
            with col1:
                st.metric("Total Transactions", f"{stats['total_transactions']:,}")
            with col2:
                st.metric("✅ Lettrés (A)", f"{stats['nb_lettres_a']:,}", 
                         delta=f"{stats['nb_lettres_a']/stats['total_transactions']*100:.1f}%")
            with col3:
                st.metric("💰 P(FA)i", f"{stats['nb_pfa']:,}",
                         delta=f"{stats['nb_pfa']/stats['total_transactions']*100:.1f}%")
            with col4:
                st.metric("⚠️ P(REG)i", f"{stats['nb_preg']:,}",
                         delta=f"{stats['nb_preg']/stats['total_transactions']*100:.1f}%")
            with col5:
                st.metric("❌ Anomalies", f"{stats['nb_non_reglee'] + stats['nb_non_lettre']:,}",
                         delta=f"{(stats['nb_non_reglee'] + stats['nb_non_lettre'])/stats['total_transactions']*100:.1f}%")
            
            # Graphique
            st.markdown("### 📈 Répartition des Codes")
            chart_data = pd.DataFrame({
                'Code': ['A', 'P(FA)i', 'P(REG)i', 'NON REGLEE', 'NON LETTRE'],
                'Nombre': [
                    stats['nb_lettres_a'],
                    stats['nb_pfa'],
                    stats['nb_preg'],
                    stats['nb_non_reglee'],
                    stats['nb_non_lettre']
                ]
            })
            st.bar_chart(chart_data.set_index('Code'))
            
            # Aperçu du tableau
            st.markdown("### 👀 Aperçu des Données Lettrées")
            st.dataframe(
                df_resultat[['Étiquettes de lignes', 'NUM_AP', 'TRANSACTION_NUMBER', 
                            'TYPE_TRANSACTION', 'SOLDE', 'LETTRAGE_AUTO']].head(20),
                use_container_width=True
            )
            
            # Téléchargement
            st.markdown("### 💾 Télécharger les Résultats")
            
            # Créer le fichier Excel en mémoire
           

            ##################
            output = BytesIO()

            with pd.ExcelWriter(output, engine='openpyxl') as writer:
    
                df_export = df_resultat.copy()
                df_export.to_excel(writer, sheet_name="Lettrage", index=False)
    
                wb = writer.book
                ws = writer.sheets["Lettrage"]
    
    # =========================
    # 🎨 Styles
    # =========================
    
                header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
                header_font = Font(color="FFFFFF", bold=True)
                header_alignment = Alignment(horizontal="center", vertical="center")
    
                color_a = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                color_non_reglee = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                color_non_lettre = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                color_pfa = PatternFill(start_color="B4C7E7", end_color="B4C7E7", fill_type="solid")
                color_preg = PatternFill(start_color="F4B084", end_color="F4B084", fill_type="solid")
    
    # =========================
    # 🔵 Colorer les en-têtes
    # =========================
    
                for col_idx, col_name in enumerate(df_export.columns, start=1):
                    cell = ws.cell(row=1, column=col_idx)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
    
    # Trouver la colonne LETTRAGE_AUTO dynamiquement
                lettrage_col_index = df_export.columns.get_loc("LETTRAGE_AUTO") + 1
    
    # =========================
    # 🎨 Colorer LETTRAGE_AUTO
    # =========================
    
                for row_idx in range(2, len(df_export) + 2):
                    cell = ws.cell(row=row_idx, column=lettrage_col_index)
                    value = cell.value
        
                    if value == 'A':
                        cell.fill = color_a
                    elif value == 'NON REGLEE':
                        cell.fill = color_non_reglee
                    elif value == 'NON LETTRE':
                        cell.fill = color_non_lettre
                    elif value and 'P(FA)' in str(value):
                        cell.fill = color_pfa
                    elif value and 'P(REG)' in str(value):
                        cell.fill = color_preg
    
    # =========================
    # 📏 Ajuster largeur colonnes
    # =========================
    
                for column_cells in ws.columns:
                    max_length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
                    ws.column_dimensions[get_column_letter(column_cells[0].column)].width = max_length + 2

            output.seek(0)



            
            st.download_button(
                label="📥 Télécharger le Fichier Lettré (Excel)",
                data=output,
                file_name=f"Lettrage_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary",
                use_container_width=True
            )
    
    # ========== ONGLET 2: STATISTIQUES ==========
    with tab2:
        if st.session_state.get('lettrage_effectue', False):
            st.markdown("## 📊 Statistiques Détaillées")
            
            stats = st.session_state['stats']
            df_resultat = st.session_state['df_resultat']
            
            # KPIs
            col1, col2 = st.columns(2)
            
            with col1:
                st.markdown("### 🎯 Indicateurs de Performance")
                taux_lettrage = stats['nb_lettres_a'] / stats['total_transactions'] * 100
                taux_qualite = (stats['nb_lettres_a'] + stats['nb_pfa'] + stats['nb_preg']) / stats['total_transactions'] * 100
                
                st.metric("Taux de Lettrage Exact", f"{taux_lettrage:.1f}%",
                         delta="✅ Excellent" if taux_lettrage > 70 else "⚠️ À améliorer")
                st.metric("Taux de Qualité Globale", f"{taux_qualite:.1f}%",
                         delta="✅ Excellent" if taux_qualite > 85 else "⚠️ À améliorer")
            
            with col2:
                st.markdown("### 💰 Montants Impliqués")
                montant_total = df_resultat['SOLDE'].abs().sum()
                montant_non_reglee = df_resultat[df_resultat['LETTRAGE_AUTO'] == 'NON REGLEE']['SOLDE'].abs().sum()
                montant_non_lettre = df_resultat[df_resultat['LETTRAGE_AUTO'] == 'NON LETTRE']['SOLDE'].abs().sum()
                
                st.metric("Total", f"{montant_total:,.2f} DT")
                st.metric("Factures Non Réglées", f"{montant_non_reglee:,.2f} DT")
                st.metric("Règlements Orphelins", f"{montant_non_lettre:,.2f} DT")
            
            # Tableaux détaillés
            st.markdown("### 📋 Analyse par Code")
            
            codes_stats = []
            for code in ['A', 'NON REGLEE', 'NON LETTRE']:
                df_code = df_resultat[df_resultat['LETTRAGE_AUTO'] == code]
                codes_stats.append({
                    'Code': code,
                    'Transactions': len(df_code),
                    'Pourcentage': f"{len(df_code)/len(df_resultat)*100:.1f}%",
                    'Montant (DT)': f"{df_code['SOLDE'].abs().sum():,.2f}"
                })
            
            # P(FA) et P(REG)
            df_pfa = df_resultat[df_resultat['LETTRAGE_AUTO'].astype(str).str.match(r'P\(FA\)\d+', na=False)]
            df_preg = df_resultat[df_resultat['LETTRAGE_AUTO'].astype(str).str.match(r'P\(REG\)\d+', na=False)]
            
            codes_stats.append({
                'Code': 'P(FA)i',
                'Transactions': len(df_pfa),
                'Pourcentage': f"{len(df_pfa)/len(df_resultat)*100:.1f}%",
                'Montant (DT)': f"{df_pfa['SOLDE'].abs().sum():,.2f}"
            })
            
            codes_stats.append({
                'Code': 'P(REG)i',
                'Transactions': len(df_preg),
                'Pourcentage': f"{len(df_preg)/len(df_resultat)*100:.1f}%",
                'Montant (DT)': f"{df_preg['SOLDE'].abs().sum():,.2f}"
            })
            
            st.dataframe(pd.DataFrame(codes_stats), use_container_width=True)
            
            # TOP Anomalies
            st.markdown("### 🚨 TOP 10 Anomalies à Traiter")
            
            col1, col2 = st.columns(2)
            
            with col1:
                st.markdown("#### Factures Non Réglées")
                df_non_reglee = df_resultat[df_resultat['LETTRAGE_AUTO'] == 'NON REGLEE'].copy()
                df_non_reglee['Montant_Abs'] = df_non_reglee['SOLDE'].abs()
                df_non_reglee = df_non_reglee.nlargest(10, 'Montant_Abs')
                st.dataframe(
                    df_non_reglee[['Étiquettes de lignes', 'NUM_AP', 'Montant_Abs']],
                    use_container_width=True
                )
            
            with col2:
                st.markdown("#### Règlements Orphelins")
                df_non_lettre = df_resultat[df_resultat['LETTRAGE_AUTO'] == 'NON LETTRE'].copy()
                df_non_lettre['Montant_Abs'] = df_non_lettre['SOLDE'].abs()
                df_non_lettre = df_non_lettre.nlargest(10, 'Montant_Abs')
                st.dataframe(
                    df_non_lettre[['Étiquettes de lignes', 'TRANSACTION_NUMBER', 'Montant_Abs']],
                    use_container_width=True
                )
        
        else:
            st.info("👆 Veuillez d'abord effectuer le lettrage dans l'onglet 'Upload & Lettrage'")
    
    # ========== ONGLET 3: DOCUMENTATION ==========
    with tab3:
        st.markdown("## 📖 Documentation")
        
        st.markdown("""
        ### 🎯 Objectif
        
        Cette application automatise le processus de lettrage comptable en établissant 
        la correspondance entre les règlements et les factures.
        
        ### 📋 Fichiers Requis
        
        1. **Règlement Global** : Fichier Excel contenant les correspondances entre :
           - Numéro de Règlement
           - Numéro AP de Facture
        
        2. **Justificatif Global** : Fichier Excel contenant toutes les transactions avec :
           - Étiquettes de lignes (Fournisseur)
           - NUM_AP (pour les factures)
           - TRANSACTION_NUMBER (pour les règlements)
           - SOLDE (montant de la transaction)
        
        ### 🏷️ Codes de Lettrage
        
        | Code | Signification | Description |
        |------|---------------|-------------|
        | **A** | Lettré Exactement | Règlements = Factures (écart < 0.001 DT) |
        | **P(FA)i** | Excédent | Règlements > Factures |
        | **P(REG)i** | Paiement Partiel | Règlements < Factures |
        | **NON REGLEE** | Non Réglée | Facture sans règlement |
        | **NON LETTRE** | Non Lettré | Règlement sans facture |
        
        ### 🔄 Processus de Lettrage
        
        1. **Chargement** : Import des deux fichiers Excel
        2. **Mapping** : Construction des correspondances règlements ↔ factures
        3. **Identification** : Détermination du type de chaque transaction (RÈGLEMENT ou FACTURE)
        4. **Groupement** : Rassemblement de toutes les transactions liées
        5. **Calcul** : Comparaison des totaux (règlements vs factures)
        6. **Attribution** : Assignation du code approprié à tout le groupe
        
        ### 💡 Logique Appliquée
        
        Pour chaque **groupe** de transactions liées :
        
        ```
        Écart = Total Règlements - Total Factures
        
        SI |Écart| < 0.001 DT  → Code "A"
        SI Écart > 0.001 DT     → Code "P(FA)i"
        SI Écart < -0.001 DT    → Code "P(REG)i"
        ```
        
        **Important** : Toutes les transactions d'un même groupe ont le **même code**.
        
        ### 📊 Indicateurs de Performance
        
        - **Taux de Lettrage Exact** : % de transactions avec code A (Cible > 70%)
        - **Taux de Qualité Globale** : % (A + P(FA) + P(REG)) (Cible > 85%)
        - **Factures Non Réglées** : % de NON REGLEE (Cible < 10%)
        - **Règlements Orphelins** : % de NON LETTRE (Cible < 5%)
        
        ### 🚀 Déploiement
        
        **Déploiement local** :
        ```bash
        streamlit run app_lettrage.py
        ```
        
        **Déploiement en ligne** :
        - Streamlit Cloud (gratuit)
        - Heroku
        - AWS / Azure / GCP
        
        ### 📞 Support
        
        Pour toute question ou problème :
        1. Vérifier la structure des fichiers d'entrée
        2. Consulter les exemples dans la documentation
        3. Contacter le support technique
        
        ---
        
        **Version** : 1.0  
        **Date** : Décembre 2025  
        **Auteur** : Système de Lettrage Automatique
        """)


# ============================================================================
# POINT D'ENTRÉE
# ============================================================================

if __name__ == "__main__":
    # Initialiser session_state si nécessaire
    if 'lettrage_effectue' not in st.session_state:
        st.session_state['lettrage_effectue'] = False
    
    main()